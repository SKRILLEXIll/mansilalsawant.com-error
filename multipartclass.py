import urllib3
import requests
from openpyxl import Workbook
from openpyxl import load_workbook


class MultiPartReq(object):
    """docstring for MulntiPartReq"""
    def makeReq(self,Q):
        with requests.Session() as s:
            p = s.post('http://mansilalsawant.com/wp-admin/admin-ajax.php',files=Q)
            return p.text  
    def makeBinaryBody(self,arr):
        Q = {
            "qsm_hidden_questions" : (None,""),
            "question11" : (None,arr[0]),
            "question12" : (None,arr[1]),
             "question13" : (None,arr[2]),
             "question14" : (None,arr[3]),
             "question15" : (None,arr[4]),
             "question16" : (None,arr[5]),
             "question17" : (None,arr[6]),
             "question18" : (None,arr[7]),
             "question19" : (None,arr[8]),
             "question20" : (None,arr[9]),
             "qmn_question_list" : (None,"11Q12Q13Q14Q15Q16Q17Q18Q19Q20Q"),
             "contact_field_0" : (None,"dsdsd"),
             "contact_field_1" : (None,"sahilpawar5566@gmail.com"),
             "contact_field_2" : (None,"4433343223"),
             "contact_field_3" : (None,"3232323"),
             "qmn_all_questions_count" : (None,"10"),
             "total_questions" : (None,"10"),
             "timer" : (None,"212"),
             "timer_ms" : (None,"1657167796376"),
             "qmn_quiz_id" : (None,"2"),
             "complete_quiz" : (None,"confirmation"),
             "action" : (None,"qmn_process_quiz"),
             "nonce" : (None,"0323f381fe"),
             "currentuserTime" : (None,"1657168009"),
             "currentuserTimeZone" : (None,"Asia/Calcutta")
        }
        return Q  
    def makeProxyReq(self):
        proxies = {
            'http': 'http://127.0.0.1:8080'
        }
        return proxies
    def strtobool(self,str):
        arr=[]
        for i in str:
            min = True if i=='1' else False
            arr.append(min)
        print(arr)
    def verify(self,str):
        if "Nonce" in str or "failed!" in str or "Validation" in str:
            return "fail"
        else:
            return "pass"
    def createWorkbook(self,filename,Q):
        path = 'E:\\pycharm\\projects\\imagebuilging\\'
        content= []    
        #content.append(Q) # adding empty cells in col 1 + 2
        wb = load_workbook(path + filename)
        ws = wb.active
        #print(type(Q))
        ws.append(Q)
        wb.save(path + filename)

r1 = MultiPartReq()
#r1.makeReq(r1.makeBinaryBody())     

for i in range(0,1024):
    arr=list('{0:010b}'.format(i))
    newarr=list()
    txt= r1.makeReq(r1.makeBinaryBody(arr))
    #print(txt)
    result=r1.verify(txt)
    newarr.append(str(i))
    newarr=newarr+ arr
    newarr.append(str(result))
    #print(newarr)
    print(str(arr)+"  ==>  "+r1.verify(txt) +" ="+str(i))
    r1.createWorkbook("tim.xlsx",newarr)

